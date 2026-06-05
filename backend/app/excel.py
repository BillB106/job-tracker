"""Build the formatted Excel export in-memory.

Single source of truth for the export layout, used by the /api/export.xlsx
endpoint (UI button) and the build_xlsx.py CLI script.

Column order mirrors the web grid; Job Title is hyperlinked; rows are
colour-coded by Prio; Offline jobs are shown italic + strikethrough + grey.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Job Portal", "job_portal"),
    ("Prio", "prio"),
    ("Status", "status"),
    ("Job Title", "title"),
    ("Company", "company"),
    ("Location", "location"),
    ("Work Type", "work_type"),
    ("Posted (approx)", "posted_date"),
    ("Posted", "posted"),
    ("Applicants / Clicks", "applicants"),
    ("Apply Method", "apply_method"),
    ("Notes", "notes"),
]
WIDTHS = [13, 18, 10, 48, 22, 26, 11, 14, 18, 20, 13, 34]
TITLE_COL = [k for _, k in COLUMNS].index("title") + 1

PRIO_FILLS = {
    "Tackle now": "FCDCD8",
    "Keep an eye on": "FDF2C9",
    "Applied / in progress": "D8F3DF",
}
GREY = "AAB2BD"


def _row_font(job, *, link=False):
    """Font for a data cell, honouring Offline (italic+strike+grey) and Discard (grey)."""
    offline = job.get("status") == "Offline"
    discard = job.get("prio") == "Discard / not interested"
    if offline:
        return Font(color=GREY, italic=True, strike=True, underline="single" if link else None)
    if discard:
        return Font(color=GREY, underline="single" if link else None)
    if link:
        return Font(color="0563C1", underline="single")
    return Font()


def build_xlsx_bytes(jobs: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Saved Jobs"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, (h, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border

    for i, job in enumerate(jobs, 1):
        excel_row = i + 1
        prio = job.get("prio", "")
        row_fill = PatternFill("solid", fgColor=PRIO_FILLS[prio]) if prio in PRIO_FILLS else None
        body_font = _row_font(job)
        for c, (_, key) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=excel_row, column=c, value=job.get(key, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (TITLE_COL, len(COLUMNS))))
            cell.font = body_font
            if row_fill:
                cell.fill = row_fill
        tc = ws.cell(row=excel_row, column=TITLE_COL)
        if job.get("source_url"):
            tc.hyperlink = job["source_url"]
        tc.font = _row_font(job, link=True)

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(jobs)+1}"
    ws.row_dimensions[1].height = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
