#!/usr/bin/env python3
"""Export the LIVE Job Tracker data to a formatted Excel file.

Pulls the current jobs from the running app (so it reflects whatever Prio /
Status / Notes you've set in the web tool), then writes an .xlsx whose column
order mirrors the web grid:
  Job Portal | Prio | Status | Job Title (hyperlinked) | Company | Location |
  Work Type | Posted (approx) | Posted | Applicants / Clicks | Apply Method | Notes

Usage:
    python3 build_xlsx.py --base-url http://127.0.0.1:8000 --password "$APP_PASSWORD"
    python3 build_xlsx.py --base-url https://YOUR-APP-URL --password "..." --out ~/jobs.xlsx
"""
import argparse
import sys
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (header, job-dict key)
COLUMNS = [
    ("Job Portal", "job_portal"),
    ("Prio", "prio"),
    ("Status", "status"),
    ("Job Title", "title"),
    ("Company", "company"),
    ("Location", "location"),
    ("Work Type", "work_type"),
    ("Posted (approx)", "posted_date"),
    ("Posted", "posted"),
    ("Applicants / Clicks", "applicants"),
    ("Apply Method", "apply_method"),
    ("Notes", "notes"),
]
WIDTHS = [13, 18, 10, 48, 22, 26, 11, 14, 18, 20, 13, 34]
TITLE_COL = [k for _, k in COLUMNS].index("title") + 1

# Optional colour coding by Prio (parity with the web tool).
PRIO_FILLS = {
    "Tackle now": "FCDCD8",
    "Keep an eye on": "FDF2C9",
    "Applied / in progress": "D8F3DF",
}
DISCARD_FONT = "AAB2BD"


def fetch_jobs(base_url, password):
    s = requests.Session()
    if s.post(f"{base_url}/api/login", json={"password": password}).status_code != 200:
        sys.exit("Login failed - check --password / --base-url.")
    r = s.get(f"{base_url}/api/jobs")
    if r.status_code != 200:
        sys.exit(f"Could not fetch jobs ({r.status_code}).")
    return r.json()


def build(jobs, out, colour=True):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saved Jobs"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font = Font(color="0563C1", underline="single")

    for c, (h, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
        cell.border = border

    for i, job in enumerate(jobs, 1):
        excel_row = i + 1
        prio = job.get("prio", "")
        row_fill = PatternFill("solid", fgColor=PRIO_FILLS[prio]) if (colour and prio in PRIO_FILLS) else None
        dim = colour and prio == "Discard / not interested"
        for c, (_, key) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=excel_row, column=c, value=job.get(key, ""))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (TITLE_COL, len(COLUMNS))))
            if row_fill:
                cell.fill = row_fill
            if dim:
                cell.font = Font(color=DISCARD_FONT)
        # Hyperlink the Job Title cell
        tc = ws.cell(row=excel_row, column=TITLE_COL)
        if job.get("source_url"):
            tc.hyperlink = job["source_url"]
            tc.font = Font(color=(DISCARD_FONT if dim else "0563C1"), underline="single")

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(jobs)+1}"
    ws.row_dimensions[1].height = 22
    wb.save(out)


def main():
    import os
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-url", default="http://127.0.0.1:8000")
    ap.add_argument("--password", required=True)
    ap.add_argument("--out", default=os.path.expanduser("~/LinkedIn_Saved_Jobs.xlsx"))
    ap.add_argument("--no-colour", action="store_true", help="Disable Prio colour coding")
    args = ap.parse_args()

    jobs = fetch_jobs(args.base_url, args.password)
    build(jobs, args.out, colour=not args.no_colour)
    print(f"Exported {len(jobs)} live jobs -> {args.out}")


if __name__ == "__main__":
    main()
