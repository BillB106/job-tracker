#!/usr/bin/env python3
"""Seed the Job Tracker DB from the existing LinkedIn_Saved_Jobs.xlsx.

Reads the spreadsheet, derives a LinkedIn external_id from the job link, and
POSTs everything to the running app's /api/import endpoint (which upserts, so
re-running is safe and preserves your prio/status/notes).

Usage:
    python3 import_excel.py \
        --xlsx ~/LinkedIn_Saved_Jobs.xlsx \
        --base-url http://localhost:8000 \
        --password "$APP_PASSWORD"
"""
import argparse
import re
import sys
import openpyxl
import requests

# Maps the spreadsheet headers to DB fields.
HEADER_MAP = {
    "Job Title": "title",
    "Company": "company",
    "Location": "location",
    "Work Type": "work_type",
    "Posted": "posted",
    "Posted (approx)": "posted_date",
    "Applicants / Clicks": "applicants",
    "Apply Method": "apply_method",
    "Notes": "notes",
    "Job Link": "source_url",
}


def extract_id(url: str) -> str:
    m = re.search(r"/jobs/view/(\d+)", url or "")
    return m.group(1) if m else ""


def read_rows(path: str):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    # Build hyperlink lookup for the "Job Link" column (cell shows "Open ↗").
    link_col = headers.index("Job Link") + 1 if "Job Link" in headers else None
    jobs = []
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, h in enumerate(headers, 1):
            field = HEADER_MAP.get(h)
            if not field:
                continue
            cell = ws.cell(r, c)
            if h == "Job Link":
                row["source_url"] = cell.hyperlink.target if cell.hyperlink else (cell.value or "")
            else:
                row[field] = cell.value if cell.value is not None else ""
        if not row.get("title"):
            continue
        row["job_portal"] = "LinkedIn"
        row["external_id"] = extract_id(row.get("source_url", ""))
        jobs.append(row)
    return jobs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--base-url", default="http://localhost:8000")
    ap.add_argument("--password", required=True)
    args = ap.parse_args()

    jobs = read_rows(args.xlsx)
    print(f"Read {len(jobs)} jobs from {args.xlsx}")

    s = requests.Session()
    login = s.post(f"{args.base_url}/api/login", json={"password": args.password})
    if login.status_code != 200:
        sys.exit(f"Login failed ({login.status_code}). Check --password.")

    res = s.post(f"{args.base_url}/api/import", json={"portal": "LinkedIn", "jobs": jobs})
    if res.status_code != 200:
        sys.exit(f"Import failed ({res.status_code}): {res.text}")
    print("Import result:", res.json())


if __name__ == "__main__":
    main()
